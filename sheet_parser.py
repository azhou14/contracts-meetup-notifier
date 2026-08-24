"""
Parses the block-structured "Class Lunch" / "Class Coffee" tabs into a
list of Event objects.

Expected shape per tab (this is NOT a normal table — it's a stack of
repeating blocks separated by blank rows):

    A1: "Tuesdays, 12:30-1:30pm Weekly."      <- tab-level info, ignored
    A2: "Sign-Ups will be posted 14 days..."  <- tab-level info, ignored
    A3: "August 18th (The Faculty Club)"      <- block header: date (+ location)
    A4: "Name"        B4: "Email Address"     <- subheader, ignored
    A5: "Shiv Soin"   B5: "shivsoin@..."      <- attendee row
    A6: ...
    A7: ...
    (blank row)
    A9: "August 25th (TBD)"                   <- next block header
    ...

A block header is any non-blank A-cell that is NOT "Name" and that
follows a blank row (or is the first content row). Location is
whatever's in the trailing parentheses; "TBD" is treated as unknown.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from typing import Optional

import openpyxl

# Tabs we care about; everything else (e.g. "Monday Snacks") is ignored.
RELEVANT_TABS = {
    "Class Lunch": "lunch",
    "Class Coffee": "coffee",
}

HEADER_RE = re.compile(
    r"^\s*(?P<month_day>[A-Za-z]+\s+\d{1,2})(?:st|nd|rd|th)?\s*"
    r"(?:\((?P<location>[^)]*)\))?\s*$"
)


@dataclass
class Attendee:
    name: str
    email: str

@dataclass
class Event:
    kind: str  # "lunch" or "coffee"
    tab_name: str
    event_date: date
    location: Optional[str]  # None if TBD / unspecified
    attendees: list[Attendee] = field(default_factory=list)
    header_row: int = 0

    @property
    def location_display(self) -> str:
        return self.location if self.location else "TBD"


def _parse_header(cell_value: str, today: date) -> Optional[tuple[date, Optional[str]]]:
    """Parse a header like 'August 25th (TBD)' -> (date, location|None).

    The sheet has no year, so we infer it: assume the *next* occurrence
    of that month/day relative to `today` (handles a Dec -> Jan rollover
    if this ever runs across a year boundary).
    """
    m = HEADER_RE.match(cell_value)
    if not m:
        return None

    month_day = m.group("month_day")
    location = m.group("location")
    if location is not None:
        location = location.strip()
        if location.upper() == "TBD" or location == "":
            location = None

    # Try this year first, then next year, and pick whichever is closest
    # to (but not absurdly far before) today.
    for year_offset in (0, 1):
        candidate_year = today.year + year_offset
        try:
            parsed = _strptime_month_day(month_day, candidate_year)
        except ValueError:
            continue
        # Allow a little slack (e.g. running the check a day late) but
        # otherwise prefer a date that is today or in the future.
        if parsed >= today.replace(day=1):  # generous lower bound
            return parsed, location

    # Fall back to current-year parse even if it's in the past, so a
    # malformed date still surfaces rather than silently vanishing.
    try:
        return _strptime_month_day(month_day, today.year), location
    except ValueError:
        return None


def _strptime_month_day(month_day: str, year: int) -> date:
    from datetime import datetime

    dt = datetime.strptime(f"{month_day} {year}", "%B %d %Y")
    return dt.date()


def parse_workbook(path: str, today: date) -> list[Event]:
    """Read the workbook and return all parsed events across relevant tabs."""
    wb = openpyxl.load_workbook(path, data_only=True)
    events: list[Event] = []

    for tab_name, kind in RELEVANT_TABS.items():
        if tab_name not in wb.sheetnames:
            continue
        ws = wb[tab_name]
        events.extend(_parse_tab(ws, tab_name, kind, today))

    return events


def _row_is_blank(ws, row: int, max_col: int = 2) -> bool:
    return all(ws.cell(row=row, column=c).value in (None, "") for c in range(1, max_col + 1))


def _parse_tab(ws, tab_name: str, kind: str, today: date) -> list[Event]:
    """
    Single pass, keyed purely on row content (not on blank-row tracking,
    which breaks for the first block since rows 1-2 are tab-level info
    and aren't blank):

      - A-cell matches the date-header pattern  -> start a new Event
      - A-cell == "Name"                          -> subheader, skip
      - A-cell blank                               -> separator, skip
      - anything else while inside a block         -> attendee row
      - anything else before any block started      -> tab-level info, skip
    """
    events: list[Event] = []
    current: Optional[Event] = None

    for row in range(1, ws.max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        b_val = ws.cell(row=row, column=2).value
        a_str = str(a_val).strip() if a_val not in (None, "") else ""

        if not a_str:
            continue  # blank separator row

        if a_str == "Name":
            continue  # subheader row

        parsed = _parse_header(a_str, today)
        if parsed:
            if current:
                events.append(current)
            event_date, location = parsed
            current = Event(
                kind=kind,
                tab_name=tab_name,
                event_date=event_date,
                location=location,
                header_row=row,
            )
            continue

        if current is not None:
            email = str(b_val).strip() if b_val not in (None, "") else ""
            if email:  # only count rows that actually have an email
                current.attendees.append(Attendee(name=a_str, email=email))
        # else: tab-level info line before the first block (e.g. row 1-2) - skip

    if current:
        events.append(current)

    return events


def events_on(events: list[Event], target_date: date) -> list[Event]:
    return [e for e in events if e.event_date == target_date]
